import base64
import logging
import os
import re
from pathlib import Path
from threading import Lock
from typing import List, Optional

from fastapi import Body, FastAPI
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook


APP_NAME = "OCR Image Receiver"

# .env 파일이 있으면 수동으로 로드
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    with open(env_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ[key.strip()] = value.strip()

# 환경변수에서 결과 디렉토리 경로를 가져오거나, 기본값으로 실행파일이 있는 디렉토리 사용
env_result_dir = os.getenv("OCR_RESULT_DIR")
print(f"[DEBUG] OCR_RESULT_DIR 환경변수: {env_result_dir}")
print(f"[DEBUG] .env 파일 경로: {env_file}")
print(f"[DEBUG] .env 파일 존재 여부: {env_file.exists()}")

RESULT_DIR = Path(os.getenv("OCR_RESULT_DIR", Path(__file__).resolve().parent)).resolve()
print(f"[DEBUG] 최종 RESULT_DIR: {RESULT_DIR}")

LOG_XLSX_PATH = Path(__file__).resolve().parent / "활동데이터_견적서템플릿.xlsx"
_EXCEL_LOCK = Lock()


class ImagePayload(BaseModel):
    filename: str
    data: str  # base64 string; may also be a data URI


class CancelRequest(BaseModel):
    reason: Optional[str] = None


app = FastAPI(title=APP_NAME)


class ResponseItem(BaseModel):
    partNumber: str          # 품번
    itemCategory: str        # 품목
    productName: str         # 품명
    specifications: str      # 규격
    quantity: str            # 수량
    unitPrice: str           # 단가
    totalAmount: str         # 최종금액
    manufacturer: str        # 제조사
    size: str                # 사이즈


def ensure_result_dir() -> None:
    RESULT_DIR.mkdir(parents=True, exist_ok=True)


_WINDOWS_RESERVED = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def sanitize_filename(original_name: str) -> str:
    # Keep only the base name and strip whitespace
    name_only = Path(original_name).name.strip()
    # Replace characters illegal on Windows
    name_only = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", name_only)
    # Avoid reserved device names
    stem = Path(name_only).stem.upper()
    if stem in _WINDOWS_RESERVED:
        name_only = f"_{name_only}"
    # Ensure not empty
    return name_only or "unnamed.png"


def decode_base64_maybe_data_uri(data: str) -> bytes:
    raw = data.strip()
    # Support data URIs like: data:image/png;base64,AAAA
    if "," in raw and ";base64" in raw.split(",", 1)[0]:
        raw = raw.split(",", 1)[1]
    # Fix missing padding
    padding_len = (-len(raw)) % 4
    if padding_len:
        raw += "=" * padding_len
    return base64.b64decode(raw, validate=False)


def save_image_to_result_dir(filename: str, content: bytes) -> Path:
    safe_name = sanitize_filename(filename)
    # Default extension to .png if absent
    if not Path(safe_name).suffix:
        safe_name = f"{safe_name}.png"
    dest = RESULT_DIR / safe_name
    with open(dest, "wb") as f:
        f.write(content)
    return dest


def handle_images(payload: List[ImagePayload]) -> List[str]:
    ensure_result_dir()
    saved_names: List[str] = []
    for item in payload:
        try:
            binary = decode_base64_maybe_data_uri(item.data)
        except Exception as e:
            logging.exception("Failed to decode base64 for %s", item.filename)
            print(f"[RECV] decode error: {item.filename}")
            saved_names.append(sanitize_filename(item.filename))
            continue

        try:
            dest = save_image_to_result_dir(item.filename, binary)
            print(f"[SAVE] {item.filename} -> {dest}")
            saved_names.append(dest.name)
        except Exception as e:
            logging.exception("Failed to save file %s", item.filename)
            print(f"[SAVE-ERROR] {item.filename}: {e}")
            saved_names.append(sanitize_filename(item.filename))
    return saved_names


def build_response_items(saved_names: List[str]) -> List[ResponseItem]:
    items: List[ResponseItem] = []
    for idx, name in enumerate(saved_names, start=1):
        items.append(
            ResponseItem(
                partNumber="partNumber",
                itemCategory="품목",
                productName=name,
                specifications="specifications",
                quantity="10",
                unitPrice="1400",
                totalAmount="14000",
                manufacturer="nextware",
                size="size",
            )
        )
    return items


def append_to_excel(items: List[ResponseItem]) -> None:
    try:
        _EXCEL_LOCK.acquire()
        if LOG_XLSX_PATH.exists():
            wb = load_workbook(LOG_XLSX_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["품번", "품목", "품명", "규격", "수량", "단가", "최종금액", "제조사", "사이즈"])
        for it in items:
            ws.append([it.partNumber, it.itemCategory, it.productName, it.specifications, 
                      it.quantity, it.unitPrice, it.totalAmount, it.manufacturer, it.size])
        wb.save(LOG_XLSX_PATH)
    except Exception:
        logging.exception("Failed to append responses to Excel: %s", LOG_XLSX_PATH)
    finally:
        try:
            _EXCEL_LOCK.release()
        except Exception:
            pass


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "result_dir": str(RESULT_DIR)}


@app.post("/ocr", response_model=List[ResponseItem])
def upload_ocr(payload: List[ImagePayload] = Body(...)) -> List[ResponseItem]:
    try:
        print(f"[RECV] /ocr items={len(payload)} files={[p.filename for p in payload]}")
    except Exception:
        print("[RECV] /ocr (unable to print filenames)")
    saved_names = handle_images(payload)
    items = build_response_items(saved_names)
    append_to_excel(items)
    return items


# Also accept POST at root for convenience
@app.post("/", response_model=List[ResponseItem])
def upload_root(payload: List[ImagePayload] = Body(...)) -> List[ResponseItem]:
    try:
        print(f"[RECV] / items={len(payload)} files={[p.filename for p in payload]}")
    except Exception:
        print("[RECV] / (unable to print filenames)")
    saved_names = handle_images(payload)
    items = build_response_items(saved_names)
    append_to_excel(items)
    return items


@app.post("/cancel", response_model=str)
def cancel(req: CancelRequest) -> str:
    reason = (req.reason or "").strip()
    logging.info("Cancel requested%s", f": {reason}" if reason else "")
    print(f"[CANCEL] reason='{reason}'")
    return "CANCELLED"


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=18333)


